import re
import openpyxl

EXPECTED_HEADERS = ["序号", "人物", "镜头", "景别", "构图", "画面内容", "时长"]


def find_header_row(ws):
    for row_idx in range(1, min(ws.max_row + 1, 10)):
        values = [str(ws.cell(row=row_idx, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
        matches = sum(1 for h in EXPECTED_HEADERS if h in values)
        if matches >= 4:
            col_map = {}
            for h in EXPECTED_HEADERS:
                if h in values:
                    col_map[h] = values.index(h) + 1
            return row_idx, col_map
    return None, None


def classify_row(row_data):
    non_empty = {k: v for k, v in row_data.items() if v and str(v).strip()}
    if not non_empty:
        return "empty"
    if len(non_empty) <= 2:
        first_val = str(list(non_empty.values())[0]).strip()
        if re.match(r'^第.+集', first_val):
            return "episode_title"
        if re.match(r'^场.+[：:]', first_val):
            return "scene_header"
    if row_data.get("画面内容") and str(row_data["画面内容"]).strip():
        return "shot"
    if row_data.get("镜头") and str(row_data["镜头"]).strip():
        return "shot"
    return "empty"


def parse_scene_header(text):
    text = str(text).strip()
    parts = re.split(r'[：:\s]+', text)
    scene = {"time": "", "space": "", "location": ""}
    for p in parts:
        p = p.strip()
        if p in ["日", "夜", "晨", "黄昏", "傍晚"]:
            scene["time"] = p
        elif p in ["内", "外", "内外"]:
            scene["space"] = p
        elif not re.match(r'^场', p) and len(p) > 0:
            scene["location"] = p
    return scene


def parse_excel(file_bytes):
    """解析上传的 Excel 字节流，返回 (workbook, rows, col_map) 或抛异常"""
    import io
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    header_row, col_map = find_header_row(ws)
    if header_row is None:
        raise ValueError("未找到表头行，请确认 Excel 包含：序号、人物、镜头、景别、构图、画面内容、时长")

    rows = []
    scene_context = {"time": "日", "space": "内", "location": ""}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_data = {}
        for field, col_idx in col_map.items():
            val = ws.cell(row=row_idx, column=col_idx).value
            row_data[field] = str(val).strip() if val is not None else ""

        row_type = classify_row(row_data)

        if row_type == "scene_header":
            first_val = next((v for v in row_data.values() if v), "")
            scene_context = parse_scene_header(first_val)

        rows.append({
            "row_idx": row_idx,
            "type": row_type,
            "data": row_data,
            "scene": dict(scene_context)
        })

    return wb, rows, col_map
