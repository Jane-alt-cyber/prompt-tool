import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def write_output(wb, rows, premium_results, opensource_results):
    """在原始 Excel 上追加4列，返回字节流"""
    ws = wb.active

    header_row = None
    for r in range(1, ws.max_row + 1):
        vals = [str(ws.cell(row=r, column=c).value or "") for c in range(1, ws.max_column + 1)]
        if "序号" in vals:
            header_row = r
            break
    if not header_row:
        header_row = 1

    base_col = ws.max_column + 1
    new_headers = [
        "优质模型-文生视频prompt",
        "开源模型-首帧文生图prompt",
        "开源模型-尾帧文生图prompt",
        "开源模型-图生视频prompt"
    ]

    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="333333")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    cell_align = Alignment(vertical="top", wrap_text=True)

    for i, h in enumerate(new_headers):
        cell = ws.cell(row=header_row, column=base_col + i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.column_dimensions[openpyxl.utils.get_column_letter(base_col)].width = 60
    ws.column_dimensions[openpyxl.utils.get_column_letter(base_col + 1)].width = 50
    ws.column_dimensions[openpyxl.utils.get_column_letter(base_col + 2)].width = 50
    ws.column_dimensions[openpyxl.utils.get_column_letter(base_col + 3)].width = 35

    for row_info in rows:
        r = row_info["row_idx"]
        if row_info["type"] != "shot":
            continue
        seq = row_info["data"].get("序号", "").strip()

        cell = ws.cell(row=r, column=base_col, value=premium_results.get(seq, ""))
        cell.alignment = cell_align
        cell.border = thin_border

        os_data = opensource_results.get(seq, {})
        for j, key in enumerate(["first_frame", "last_frame", "video_motion"]):
            cell = ws.cell(row=r, column=base_col + 1 + j, value=os_data.get(key, ""))
            cell.alignment = cell_align
            cell.border = thin_border

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
